"""Add evidence-based company personalization to an XLSX/CSV lead list.

The script fetches only the company's own website and records the source URL.
It never asks an LLM to fill missing facts: on failure it writes a review flag.

Usage:
    python personalize_companies.py input.xlsx output.xlsx
    python personalize_companies.py input.csv output.xlsx
"""

from __future__ import annotations

import re
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


UA = "Mozilla/5.0 (compatible; Polza-personalization-research/1.0)"
SPACE = re.compile(r"\s+")


def normalize_url(value: str) -> str:
    value = (value or "").strip()
    if not value.startswith(("http://", "https://")):
        value = "https://" + value
    return value


def official_fact(url: str) -> tuple[str, str, str]:
    """Return (fact, final_url, status), using homepage title/description only."""
    url = normalize_url(url)
    try:
        response = requests.get(url, timeout=15, headers={"User-Agent": UA})
        response.raise_for_status()
        response.encoding = response.apparent_encoding or response.encoding
    except requests.RequestException as exc:
        return "", url, f"НУЖНА РУЧНАЯ ПРОВЕРКА: {type(exc).__name__}"

    soup = BeautifulSoup(response.text, "html.parser")
    description = soup.find("meta", attrs={"name": re.compile("^description$", re.I)})
    text = description.get("content", "") if description else ""
    if not text and soup.title:
        text = soup.title.get_text(" ", strip=True)
    text = SPACE.sub(" ", text).strip(" .")
    if len(text) < 25:
        return "", response.url, "НУЖНА РУЧНАЯ ПРОВЕРКА: на странице нет содержательного описания"

    # Keep the fact source-bound and short. No unsupported inference is added.
    fact = text[:360].rstrip(" ,;:-")
    if fact[-1:] not in ".!?":
        fact += "."
    return fact, response.url, "Автоматически извлечено с официального сайта; проверить формулировку"


def read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        import csv
        with path.open(encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    ws = load_workbook(path, data_only=True).active
    headers = [str(c.value or "").strip() for c in ws[1]]
    return [dict(zip(headers, (str(v or "") for v in row))) for row in ws.iter_rows(min_row=2, values_only=True)]


def pick(row: dict[str, str], *names: str) -> str:
    lower = {key.lower(): value for key, value in row.items()}
    for name in names:
        if name.lower() in lower:
            return lower[name.lower()]
    return ""


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    source, target = map(Path, sys.argv[1:])
    rows = read_rows(source)
    wb = Workbook()
    ws = wb.active
    ws.title = "Персонализация"
    headers = list(rows[0]) if rows else ["Компания", "Сайт"]
    ws.append(headers + ["Персонализация", "Источник персонализации", "Статус"])
    sites = [pick(row, "Сайт", "Website", "Домен", "URL") for row in rows]
    with ThreadPoolExecutor(max_workers=10) as pool:
        results = list(pool.map(official_fact, sites))
    for row, (fact, source_url, status) in zip(rows, results):
        ws.append([row.get(h, "") for h in headers] + [fact, source_url, status])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(target)
    print(f"Saved {len(rows)} rows to {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
